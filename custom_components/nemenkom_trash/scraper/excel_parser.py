"""
Excel parser for nemenkom.lt household waste timetables.

Confirmed structure (inspected 2026 Jun–Dec workbook):
  Single sheet named "2026 m" (year in sheet name)
  Row 2: headers — A=Seniūnija, B=Kaimai, C=Gatvė, D=Savaitės diena,
                    E=Birželis, F=Liepa, G=Rugpjūtis, H=Rugsėjis,
                    I=Spalis, J=Lapkritis, K=Gruodis
  Data rows: C holds the street name (may contain many streets in one cell);
             month columns (E-K) contain strings like "14d., 28d.," or "2 d., 16 d., 30 d."
             — NOT actual Excel date objects.
"""

import io
import re
import logging
from datetime import date

import openpyxl

logger = logging.getLogger(__name__)

HEADER_TO_MONTH: dict[str, int] = {
    "sausis": 1, "vasaris": 2, "kovas": 3, "balandis": 4,
    "gegužė": 5, "birželis": 6, "liepa": 7, "rugpjūtis": 8,
    "rugsėjis": 9, "spalis": 10, "lapkritis": 11, "gruodis": 12,
}

# Matches one or more day numbers from a cell like "14d., 28d.," or "2 d., 16 d., 30 d."
DAY_PATTERN = re.compile(r"\b(\d{1,2})\s*d\.?")


def _build_col_month_map(row) -> dict[int, int]:
    col_months: dict[int, int] = {}
    for cell in row:
        if not cell.value:
            continue
        text = str(cell.value).strip().lower()
        for name, month_num in HEADER_TO_MONTH.items():
            if name in text:
                col_months[cell.column] = month_num
                break
    return col_months


def _parse_days_from_string(value) -> list[int]:
    """Extract all day numbers from a string like '14d., 28d.,' → [14, 28]."""
    if not value:
        return []
    days = []
    for m in DAY_PATTERN.finditer(str(value)):
        day = int(m.group(1))
        if 1 <= day <= 31:
            days.append(day)
    return days


def extract_dates_from_excel(
    excel_bytes: bytes,
    street_aliases: list[str],
) -> list[date]:
    """Extract all collection dates for the given street from an Excel timetable."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    collected: list[date] = []
    aliases_lower = [a.lower() for a in street_aliases]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract year from the sheet name (e.g. "2026 m" → 2026)
        year_match = re.search(r"(\d{4})", sheet_name)
        context_year = int(year_match.group(1)) if year_match else date.today().year

        # Find the header row — first row that contains month names
        col_months: dict[int, int] = {}
        header_row_idx = 0
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            col_months = _build_col_month_map(row)
            if col_months:
                header_row_idx = row_idx
                logger.debug("Sheet %s: header at row %d, months: %s", sheet_name, row_idx, col_months)
                break

        if not col_months:
            logger.debug("Sheet %s: no month headers found, skipping", sheet_name)
            continue

        # Scan data rows for the street name
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            street_found = False
            for cell in row:
                if cell.value and any(a in str(cell.value).lower() for a in aliases_lower):
                    street_found = True
                    logger.debug("Matched street at %s: %s", cell.coordinate, str(cell.value)[:80])
                    break

            if not street_found:
                continue

            for cell in row:
                month = col_months.get(cell.column)
                if month is None:
                    continue
                for day in _parse_days_from_string(cell.value):
                    try:
                        collected.append(date(context_year, month, day))
                    except ValueError:
                        pass

    result = sorted(set(collected))
    logger.debug("Excel extracted %d dates", len(result))
    return result
